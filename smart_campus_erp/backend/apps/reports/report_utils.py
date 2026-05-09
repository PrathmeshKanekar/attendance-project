"""
Report generation utilities.
Uses ReportLab for PDF and openpyxl for Excel.
All free, open-source libraries. No paid services.
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any

# ── PDF generation (ReportLab) ─────────────────────────────
def generate_attendance_pdf(
    report_data : List[Dict[str, Any]],
    title       : str,
    subtitle    : str,
    college_name: str,
    threshold   : float = 75.0,
) -> bytes:
    """
    Generate a professional attendance report PDF.
    Returns raw bytes of the PDF file.

    report_data: list of dicts with keys:
        student_name, prn, total_sessions,
        present, absent, percentage
    """
    from reportlab.lib           import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units     import cm
    from reportlab.platypus      import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize    = A4,
        rightMargin = 2 * cm,
        leftMargin  = 2 * cm,
        topMargin   = 2 * cm,
        bottomMargin= 2 * cm,
    )

    styles = getSampleStyleSheet()

    # ── Custom styles ──────────────────────────────────────
    title_style = ParagraphStyle(
        'CustomTitle',
        parent    = styles['Title'],
        fontSize  = 18,
        textColor = colors.HexColor('#1E3A5F'),
        spaceAfter= 6,
        alignment = TA_CENTER,
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent    = styles['Normal'],
        fontSize  = 11,
        textColor = colors.HexColor('#64748B'),
        spaceAfter= 4,
        alignment = TA_CENTER,
    )
    college_style = ParagraphStyle(
        'College',
        parent    = styles['Normal'],
        fontSize  = 13,
        textColor = colors.HexColor('#2563EB'),
        spaceBefore=4,
        spaceAfter= 2,
        alignment = TA_CENTER,
        fontName  = 'Helvetica-Bold',
    )
    meta_style = ParagraphStyle(
        'Meta',
        parent    = styles['Normal'],
        fontSize  = 9,
        textColor = colors.HexColor('#94A3B8'),
        alignment = TA_RIGHT,
    )

    # ── Color constants ────────────────────────────────────
    NAVY      = colors.HexColor('#1E3A5F')
    BLUE      = colors.HexColor('#2563EB')
    GREEN     = colors.HexColor('#16A34A')
    GREEN_BG  = colors.HexColor('#F0FDF4')
    RED       = colors.HexColor('#DC2626')
    RED_BG    = colors.HexColor('#FEF2F2')
    ORANGE    = colors.HexColor('#EA580C')
    GRAY_BG   = colors.HexColor('#F8FAFC')
    HEADER_BG = colors.HexColor('#1E3A5F')
    WHITE     = colors.white
    LIGHT     = colors.HexColor('#E2E8F0')

    story = []

    # ── Header ─────────────────────────────────────────────
    story.append(Paragraph(college_name, college_style))
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(subtitle, subtitle_style))
    story.append(Paragraph(
        f'Generated: {datetime.now().strftime("%d %B %Y, %I:%M %p")}',
        meta_style,
    ))
    story.append(HRFlowable(
        width='100%', thickness=2, color=BLUE, spaceAfter=12,
    ))

    # ── Summary stats ──────────────────────────────────────
    total       = len(report_data)
    below_thresh = sum(
        1 for r in report_data if r.get('percentage', 0) < threshold
    )
    above_thresh = total - below_thresh

    summary_data = [
        ['Total Students', 'Above Threshold', 'Below Threshold', 'Threshold'],
        [
            str(total),
            str(above_thresh),
            str(below_thresh),
            f'{threshold}%',
        ],
    ]
    summary_table = Table(
        summary_data,
        colWidths=[4.3 * cm, 4.3 * cm, 4.3 * cm, 4.3 * cm],
    )
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, 1), GRAY_BG),
        ('FONTNAME',   (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 1), (-1, 1), 14),
        ('TEXTCOLOR',  (1, 1), (1, 1), GREEN),
        ('TEXTCOLOR',  (2, 1), (2, 1), RED),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, 1), [GRAY_BG]),
        ('BOX',        (0, 0), (-1, -1), 1, LIGHT),
        ('INNERGRID',  (0, 0), (-1, -1), 0.5, LIGHT),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    # ── Main attendance table ──────────────────────────────
    table_header = [
        'SR', 'Student Name', 'PRN',
        'Total', 'Present', 'Absent', 'Attendance %',
    ]
    col_widths = [1 * cm, 5.5 * cm, 3 * cm,
                  1.8 * cm, 2 * cm, 2 * cm, 2.9 * cm]

    table_data = [table_header]
    for i, row in enumerate(report_data, 1):
        pct  = float(row.get('percentage', 0))
        pct_text = f'{pct:.1f}%'
        table_data.append([
            str(i),
            row.get('student_name', ''),
            row.get('prn', ''),
            str(row.get('total_sessions', 0)),
            str(row.get('present', 0)),
            str(row.get('absent', 0)),
            pct_text,
        ])

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Build row-by-row styles for color coding
    style_cmds = [
        # Header
        ('BACKGROUND',  (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR',   (0, 0), (-1, 0), WHITE),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 9),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',    (0, 1), (-1, -1), 9),
        ('BOX',         (0, 0), (-1, -1), 0.5, LIGHT),
        ('INNERGRID',   (0, 0), (-1, -1), 0.3, LIGHT),
        ('TOPPADDING',  (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 6),
        # Left-align student name
        ('ALIGN',       (1, 0), (2, -1), 'LEFT'),
    ]

    for i, row in enumerate(report_data, 1):
        pct = float(row.get('percentage', 0))
        if pct < threshold:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), RED_BG))
            style_cmds.append(('TEXTCOLOR',  (6, i), (6, i), RED))
            style_cmds.append(('FONTNAME',   (6, i), (6, i), 'Helvetica-Bold'))
        elif i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), GRAY_BG))

    main_table.setStyle(TableStyle(style_cmds))
    story.append(main_table)

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f'<font color="#DC2626">■</font> Red rows indicate attendance below '
        f'{threshold}% threshold.',
        ParagraphStyle(
            'Legend', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#64748B'),
        ),
    ))

    doc.build(story)
    return buffer.getvalue()


# ── Excel generation (openpyxl) ────────────────────────────
def generate_attendance_excel(
    report_data : List[Dict[str, Any]],
    title       : str,
    subject_name: str,
    college_name: str,
    threshold   : float = 75.0,
) -> bytes:
    """
    Generate a professional attendance Excel report.
    Returns raw bytes of the .xlsx file.
    """
    from openpyxl                          import Workbook
    from openpyxl.styles                   import (
        Font, PatternFill, Alignment,
        Border, Side, numbers,
    )
    from openpyxl.utils                    import get_column_letter
    from openpyxl.chart                    import BarChart, Reference
    from openpyxl.chart.series             import DataPoint

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    # ── Color fills ────────────────────────────────────────
    NAVY_FILL    = PatternFill('solid', fgColor='1E3A5F')
    BLUE_FILL    = PatternFill('solid', fgColor='2563EB')
    GREEN_FILL   = PatternFill('solid', fgColor='DCFCE7')
    RED_FILL     = PatternFill('solid', fgColor='FEE2E2')
    ORANGE_FILL  = PatternFill('solid', fgColor='FEF9C3')
    GRAY_FILL    = PatternFill('solid', fgColor='F8FAFC')
    WHITE_FILL   = PatternFill('solid', fgColor='FFFFFF')

    thin = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_font(color='FFFFFF', size=11, bold=True):
        return Font(name='Calibri', size=size, bold=bold, color=color)

    def center_align(wrap=False):
        return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

    # ── Title rows ─────────────────────────────────────────
    ws.merge_cells('A1:G1')
    ws['A1'] = college_name
    ws['A1'].font      = header_font(color='1E3A5F', size=14)
    ws['A1'].alignment = center_align()

    ws.merge_cells('A2:G2')
    ws['A2'] = title
    ws['A2'].font      = header_font(color='2563EB', size=12)
    ws['A2'].alignment = center_align()

    ws.merge_cells('A3:G3')
    ws['A3'] = f'Subject: {subject_name}'
    ws['A3'].font      = header_font(color='374151', size=10, bold=False)
    ws['A3'].alignment = center_align()

    ws.merge_cells('A4:G4')
    ws['A4'] = f'Generated: {datetime.now().strftime("%d %B %Y, %I:%M %p")}'
    ws['A4'].font      = header_font(color='94A3B8', size=9, bold=False)
    ws['A4'].alignment = center_align()

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 14

    # ── Summary row ────────────────────────────────────────
    ws.merge_cells('A5:G5')
    ws['A5'] = ''

    summary_row = 6
    for col, (label, val) in enumerate([
        ('Total', str(len(report_data))),
        ('Present >', f'{threshold}%',),
        ('Below Threshold', str(sum(
            1 for r in report_data if r.get('percentage', 0) < threshold
        ))),
    ], 1):
        ws.cell(summary_row, col * 2 - 1, label).font = header_font(
            color='64748B', size=9, bold=True
        )
        ws.cell(summary_row, col * 2, val).font = header_font(
            color='1E3A5F', size=11
        )

    ws.row_dimensions[summary_row].height = 18

    # ── Column headers ─────────────────────────────────────
    header_row = 8
    headers    = [
        'SR No.', 'Student Name', 'PRN Number',
        'Total Sessions', 'Present', 'Absent', 'Attendance %',
    ]
    col_widths = [7, 28, 16, 15, 10, 10, 14]

    for col_idx, (header, width) in enumerate(
        zip(headers, col_widths), 1
    ):
        cell = ws.cell(header_row, col_idx, header)
        cell.font      = header_font()
        cell.fill      = NAVY_FILL
        cell.alignment = center_align()
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 20

    # ── Data rows ──────────────────────────────────────────
    for i, row in enumerate(report_data):
        data_row = header_row + 1 + i
        pct = float(row.get('percentage', 0))

        row_fill = RED_FILL if pct < threshold else (
            GRAY_FILL if i % 2 == 0 else WHITE_FILL
        )

        row_data = [
            i + 1,
            row.get('student_name', ''),
            row.get('prn', ''),
            row.get('total_sessions', 0),
            row.get('present', 0),
            row.get('absent', 0),
            pct / 100,  # will format as percentage
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell           = ws.cell(data_row, col_idx, value)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = center_align()
            cell.font      = Font(name='Calibri', size=10)

            if col_idx == 2:  # Student name left-align
                cell.alignment = Alignment(
                    horizontal='left', vertical='center'
                )
            if col_idx == 7:  # Percentage formatting
                cell.number_format = '0.0%'
                if pct < threshold:
                    cell.font = Font(
                        name='Calibri', size=10,
                        bold=True, color='DC2626',
                    )
                elif pct >= 90:
                    cell.font = Font(
                        name='Calibri', size=10,
                        bold=True, color='16A34A',
                    )

        ws.row_dimensions[data_row].height = 16

    # ── Freeze header row ──────────────────────────────────
    ws.freeze_panes = f'A{header_row + 1}'

    # ── Auto filter ────────────────────────────────────────
    ws.auto_filter.ref = (
        f'A{header_row}:G{header_row + len(report_data)}'
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
